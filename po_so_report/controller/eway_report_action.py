from collections import defaultdict
from odoo import http
from odoo.http import request
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import io
import re


class EwayReportController(http.Controller):

    @http.route('/eway_report/export_excel', type='http', auth='user')
    def export_excel(self, **kwargs):
        date_from = kwargs.get('date_from')
        date_to = kwargs.get('date_to')

        if not (date_from and date_to):
            return request.make_response("Missing date_from or date_to parameter")

        try:
            date_from_dt = datetime.strptime(date_from, "%Y-%m-%d")
            date_to_dt = datetime.strptime(date_to, "%Y-%m-%d")
        except Exception:
            return request.make_response("Invalid date format, expected YYYY-MM-DD")

        company = request.env.company
        Payment = request.env['account.payment'].sudo()

        # Step 1: Get payments
        payments = Payment.search([
            ('company_id', '=', company.id),
            ('date', '>=', date_from_dt),
            ('date', '<=', date_to_dt),
            ('partner_type', '=', 'customer'),
            ('state', '=', 'posted')
        ])

        if not payments:
            return request.make_response("No customer payments found for the given period.")

        # Step 2: Map payments → invoices
        payment_invoice_map = defaultdict(list)
        for pay in payments:
            for inv in pay.reconciled_invoice_ids:
                payment_invoice_map[pay.id].append(inv)

        if not payment_invoice_map:
            return request.make_response("No linked invoices found for the given payments.")

        # Step 3: Get sale orders
        sale_order_names = set()
        for invoices in payment_invoice_map.values():
            for inv in invoices:
                if inv.invoice_origin:
                    sale_order_names.add(inv.invoice_origin)

        SaleOrder = request.env['sale.order'].sudo()
        sale_orders = SaleOrder.search_read(
            [('name', 'in', list(sale_order_names))],
            ['id', 'name', 'user_id', 'partner_id']
        )
        sale_order_map = {so['name']: so for so in sale_orders}
        sale_order_ids = [so['id'] for so in sale_orders]

        # Step 4: EWAY operations
        Eway = request.env['eway.operation'].sudo()
        eway_data = Eway.search_read(
            [('sale_id', 'in', sale_order_ids)],
            ['id', 'name', 'trip_start_date', 'trip_end_date', 'invoice_amount', 'sale_id']
        )
        eway_map = {e['sale_id'][0]: e for e in eway_data}

        # Step 5: Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "EWAY Report"

        bold_font = Font(bold=True)
        left_align = Alignment(horizontal='left', vertical='center')

        headers = [
            "Customer Name", "Salesperson", "EWAY Number", "Trip Start", "Trip End", "EWAY Amount",
            "Invoice Number", "Invoice Date", "Invoice Amount Untaxed",
            "Partial Payment (Actual)", "Balance Amount (Untaxed)", "Actual Amount Due",
            "Payment Date", "Payment Amount"
        ]
        ws.append(headers)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = bold_font
            cell.alignment = left_align

        row = 2

        # Step 6: Process payments & invoices
        for pay in payments:
            linked_invoices = payment_invoice_map.get(pay.id, [])
            if not linked_invoices:
                continue

            for invoice in linked_invoices:

                sale = sale_order_map.get(invoice.invoice_origin)
                eway = eway_map.get(sale['id']) if sale else None

                customer_name = invoice.partner_id.name if invoice.partner_id else ''
                salesperson = request.env['res.users'].sudo().browse(sale['user_id'][0]).name if sale and sale.get(
                    'user_id') else ''
                trip_start = eway.get('trip_start_date').strftime("%m/%d/%Y") if eway and eway.get(
                    'trip_start_date') else ''
                trip_end = eway.get('trip_end_date').strftime("%m/%d/%Y") if eway and eway.get('trip_end_date') else ''

                invoice_number = invoice.name or ''
                invoice_date = invoice.invoice_date.strftime("%m/%d/%Y") if invoice.invoice_date else ''

                invoice_untaxed = invoice.amount_untaxed
                ratio = invoice.amount_untaxed / invoice.amount_total if invoice.amount_total else 1

                # Get each partial payment applied to this invoice
                allocation_list = []
                for line in invoice.line_ids:
                    for match in line.matched_credit_ids:
                        if match.credit_move_id.payment_id:
                            allocation_list.append({
                                "pay": match.credit_move_id.payment_id,
                                "amount_paid_total": match.amount
                            })

                # running balance untaxed
                balance_untaxed = invoice_untaxed

                for alloc in allocation_list:

                    payment_untaxed = alloc["amount_paid_total"] * ratio
                    payment_untaxed = round(payment_untaxed, 2)

                    balance_untaxed = round(balance_untaxed - payment_untaxed, 2)
                    if balance_untaxed < 0:
                        balance_untaxed = 0

                    payment_date = alloc["pay"].date.strftime("%m/%d/%Y") if alloc["pay"].date else ''

                    # Partial Payment column → 0 if fully paid, else actual allocation
                    partial_payment_actual = 0 if abs(invoice.amount_residual) < 0.01 else alloc["amount_paid_total"]

                    # Loop invoice lines
                    for line in invoice.invoice_line_ids:
                        eway_number = line.eway_operation.name if line.eway_operation else ''
                        eway_amount = line.price_unit if line.price_unit else 0.0

                        row_data = [
                            customer_name, salesperson, eway_number, trip_start, trip_end, eway_amount,
                            invoice_number, invoice_date, invoice_untaxed,
                            partial_payment_actual,  # 0 if fully paid
                            balance_untaxed,
                            invoice.amount_residual,  # actual amount due
                            payment_date, alloc["pay"].amount  # Always take from payment record
                        ]

                        ws.append(row_data)

                        for col_idx in range(1, len(row_data) + 1):
                            ws.cell(row=row, column=col_idx).alignment = left_align

                        row += 1

        # auto width
        for i, col in enumerate(ws.columns, start=1):
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(i)].width = max_len + 5

        fp = io.BytesIO()
        wb.save(fp)
        fp.seek(0)

        filename = f"EWAY_Report_{date_from}_to_{date_to}.xlsx"
        return request.make_response(
            fp.getvalue(),
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', f'attachment; filename={filename}')
            ]
        )














